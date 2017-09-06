import csv
import openpyxl
from openpyxl import Workbook
from sklearn.neighbors import NearestNeighbors
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import pairwise_distances
from sklearn.metrics.pairwise import euclidean_distances
from scipy.optimize import curve_fit
from sklearn.mixture import GMM
from sklearn.mixture import GaussianMixture
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
import argparse
import cPickle as pickle


parser = argparse.ArgumentParser(description='Classifier for Burkitt lymphoma and Diffuse L. B-Cell Lymphoma')
parser.add_argument('-e', type=float, nargs = 1, help = 'epsilon value for error bound between radii', required = False)
parser.add_argument('-fBL', type=str, nargs = 1, help ='Burkkitt lymphoma training data file path', required = False)
parser.add_argument('-fDL', type=str, nargs = 1, help ='Diffuse large B-cell lymphoma training data file path', required = False)
parser.add_argument('-analysis', type=str, nargs = 1, required = True, help='denote -nDistDist for computing neichborDistanceDistribution and nucMorph for computing nucleiMorphology')
args = parser.parse_args()

#
# Given an epsilon value and file path nucleiMorphology labels
# observed measurements of area and perimeter of nuclei as
# abnormally shaped if the implied radius given by perimeter
# and area measurements are within epsilon of each other when
# assumed circular. Using kNN points are then clustered by having
# similar area/perimeter values.
#
# Example call from terminal:
# python nucleimorphology.py -e 0.2 -f './BL results.xlsx' -analysis nucMorph
def nucleiMorphology(epsilon, excelFile):
    
    results = openpyxl.load_workbook(filename = excelFile)
    sheetNames = results.get_sheet_names()
    data = results.get_sheet_by_name('Sheet1')

    featuresRaw = data.rows[0]
    attributes = []
    for cell in featuresRaw:
        attributes.append(cell.value)
    #print(attributes)


    #print(len(features))#26
    # [u'Image Location', u'Analysis Region', u'Analysis Inputs', u'Object Id', u'XMin', u'XMax', u'YMin', u'YMax', u'Stain 1 Positive', u'Stain 1 Weak Positive', u'Stain 1 Moderate Positive', u'Stain 1 Strong Positive', u'Stain 2 Positive', u'Stain 2 Weak Positive', u'Stain 2 Moderate Positive', u'Stain 2 Strong Positive', u'Dual Positive', u'Neutral', u'Stain 1 Nucleus OD', u'Stain 2 Nucleus OD', u'Stain 1 Cytoplasm OD', u'Stain 2 Cytoplasm OD', u'Cell Area', u'Nucleus Area', u'Cytoplasm Area', u'Nucleus Perimeter (um)']
    # get all nuclei area and perimeter
    nucArea = []
    nucPerim = []
    ratAP = []
    approxAreaRad = []
    approxPerimRad = []
    areaPerim = []
    approxRadLabel = []
    Xmin, Xmax, Ymin, Ymax = [], [], [], []
    
    for row in range(2, data.get_highest_row()-1):
        #Collect nuclei perimeter, area, and their ratio
        nucAr = data['X'+str(row)].value
        nucPer = data['Z'+str(row)].value
        areaPerim.append([nucAr, nucPer])
        # observed area = nucAr; observed perimeter = nucPer:
        # nucAr / nucPer ~ (pi r^2) / (2 pi r) = r/2 ;
        # r ~ 2 (nucAr / nucPer)
        apratio = float(nucAr)/float(nucPer)
        nucArea.append([nucAr])
        nucPerim.append(nucPer)
        ratAP.append([apratio])
        # observed area = nucAr : nucAr = pi r^2 -> r = (nucAr/pi)^(1/2)
        arad =  np.sqrt(float(nucAr)/np.pi)
        approxAreaRad.append( arad )
        # observed perimeter = nucPer: nucPer = 2 pi r -> r = nucPer/(2 pi)
        prad =  nucPer/(2.*np.pi)
        approxPerimRad.append( nucPer/(2.*np.pi) )

        # add label 1 if radii are approximately equal
        if min(arad, prad) >= max(arad, prad) - epsilon:
            approxRadLabel.append(1)
        else:
            approxRadLabel.append(0)
            
        #Collect Image location, Xmin, Xmax, Ymin, Ymax
        Xmin.append(data['E'+str(row)].value)
        Xmax.append(data['F'+str(row)].value)
        Ymin.append(data['G'+str(row)].value)
        Ymax.append(data['H'+str(row)].value)
    #
    # Plot histogram relating radius implied by measured area and perimeter
    #
    """
    bins = np.linspace(0, 8, 200)
    plt.hist(approxAreaRad, bins, alpha= 0.5, label='Approx Radius (Area)')
    plt.hist(approxPerimRad, bins, alpha=0.5, label='Approx Radius (Perim)')
    plt.legend(loc='upper right')
    plt.title('BL approx radius w/r/t observed perim, area')
    plt.show()
    """

    #
    # Printing block for examples
    #
    #print("nuclei Area ex: %f  nuclei Perim ex: %f  area/perim ex: %f  approx area radius ex: %f approx perim radius ex: %f" % (nucArea[1:4], nucPerim[1:4], map(lambda x: float(x[0])*2., ratAP)[1:10], approxAreaRad[1:10], approxPerimRad[1:10]) )
   


    
    
    #
    # Nearest neighbor by distance for area perimeter ratios, no labeling
    # use index of closest point to find related label
    #
    """
    APmap = NearestNeighbors(n_neighbors = 1)
    #APmap.fit(np.array(ratAP))
    APmap.fit(areaPerim)
    closest = APmap.kneighbors([[18.2892, 23.184]], return_distance=False)
    #print(closest)
    #print(approxRadLabel[closest[0][0]])
    """

    #
    # K nearest neighbors. Points are taken in R^2 as (area, perimeter)
    # points considered near w/r/t to euclidean distance are then
    # similar in both area and perimeter size. Clustering is then
    # taken as nuclei similar in area/perimeter allowing unseen examples
    # to be labeled similarly.
    #
    kneigh = KNeighborsClassifier(n_neighbors = 10)
    kneigh.fit(areaPerim, approxRadLabel)
    #print("prediction with rand (area,perim) ")
    #print(kneigh.predict([[23.4, 600], [15.1, 13.7]])) # epsilon = 0.4 given [2.4, 3.2] predicts label 1, i.e. non lymphoma
    

#
# neighborDistanceDistribution() computes the midpoint of the bounding box
# of each nuclei as taken from the excell sheets BLFile and DLFile. Two methods contained
# in neighborDistanceDistribution are fitPairwiseDist(pwDist) which takes
# a pairwise distance matrix, computes the historgram for those distances and
# fits a curve to the midpoint of each bin. The second method (which for the time
# being is being called and which is still being worked on) is
# fitPairwiseDistGMM(pwDistBL, pwDistDL) which given the pairwise distance matrices
# for the Burkettes lymphoma (BL) and DLBCL (DL) trains a Gaussian mixture model
# classifier on each distribution, allowing us to later attribute new samples as
# being more likely to belong to one or the other distribution.
#
# example call:
# python lymphomaClassification.py -analysis nDistDist -fDL '../data/DLBCL results.xlsx' -fBL '../data/BL results.xlsx'
def neighborDistanceDistribution(BLFile, DLFile):

    # Since data is so large for the mean time use cPickle to
    # serialize data and load when needed.
    try:
        BLpwDist = pickle.load(open("/Users/multivax/Documents/PhD/Research/Pascucci/ML_Lymphoma/pairwiseDistBL.p", "rb"))
        DLpwDist = pickle.load(open("/Users/multivax/Documents/PhD/Research/Pascucci/ML_Lymphoma/pairwiseDistDL.p", "rb"))

        #
        # Data still to large so truncating for the time being
        #
        BLpwDist = BLpwDist[:500]
        DLpwDist = DLpwDist[:500]
        print("read data from pre-processed pickle files")
        
    except  (OSError, IOError, EOFError, StandardError) as e:
        print("No pre-processed data files. Computing mid points of nuclei and pairwise distance matrices")
        #DLBCL data
        DLresults = openpyxl.load_workbook(filename = DLFile)
        DLsheetNames = DLresults.get_sheet_names()
        DLdata = DLresults.get_sheet_by_name('Sheet1')
        
        DLfeaturesRaw = DLdata.rows[0]
        DLattributes = []
        for cell in DLfeaturesRaw:
            DLattributes.append(cell.value)
            
        #Burkitts data
        BLresults = openpyxl.load_workbook(filename = BLFile)
        BLsheetNames = BLresults.get_sheet_names()
        BLdata = BLresults.get_sheet_by_name('Sheet1')
    
        BLfeaturesRaw = BLdata.rows[0]
        BLattributes = []
        for cell in BLfeaturesRaw:
            BLattributes.append(cell.value)
        
        # compute and collect midpoints of nuclei bounding boxes
        DLnucleiMid = []
        BLnucleiMid = []
        case = -1
        for data in [DLdata, BLdata]:
            case += 1
            for row in range(2, data.get_highest_row()-1):
                #Collect Image location, Xmin, Xmax, Ymin, Ymax
                xmin, xmax, ymin, ymax = 0, 0, 0, 0
                xmin = data['E'+str(row)].value
                xmax = data['F'+str(row)].value
                ymin = data['G'+str(row)].value
                ymax = data['H'+str(row)].value
                
                mid = np.array([float(xmin+xmax)/2.,float(ymin+ymax)/2.])
                if case == 0:
                    DLnucleiMid.append(mid)
                elif case == 1:
                    BLnucleiMid.append(mid)
                

        #
        # BL is to large I run out of memory allocation so I truncate
        # to the same size as DL i.e. 40139 and 19050.
        # and much larger for pairwise distance
        #
        if len(BLnucleiMid) > 500:
            BLnucleiMid = BLnucleiMid[:500]
        if len(DLnucleiMid) > 500:
            DLnucleiMid = DLnucleiMid[:500]
        
        #
        # Compute pairwise distance matrix.(both methods use euclidean metric)
        #
        #pwDist = pairwise_distances(np.array(nucleiMid)) #same output, dif performance
        DLpwDist = euclidean_distances(np.array(DLnucleiMid),np.array(DLnucleiMid))
        BLpwDist = euclidean_distances(np.array(BLnucleiMid),np.array(BLnucleiMid))
        pDLdump = DLpwDist.flatten().tolist()
        pickle.dump(pDLdump, open('pairwiseDistDL.p', 'wb'))
        pBLdump = BLpwDist.flatten().tolist()
        pickle.dump(pBLdump, open("pairwiseDistBL.p", "wb"))
    
    #
    # Estimate distribution with histogram
    #
    def fitPairwiseDist(pwDist):

        hist, bin_edges = np.histogram(pwDist.flatten()[0], density=True)
        bin_centres = (bin_edges[:-1] + bin_edges[1:])/2

        # Define model function to be used to fit to the data above:
        def gauss(x, *p):
            A, mu, sigma = p
            return A*np.exp(-(x-mu)**2/(2.*sigma**2))

        # p0 is the initial guess for the fitting coefficients
        #(A, mu and sigma above)
        p0 = [1., 0., 1.]
        
        coeff, var_matrix = curve_fit(gauss, bin_centres, hist, p0=p0)
        
        # Get the fitted curve
        hist_fit = gauss(bin_centres, *coeff)
        
        plt.plot(bin_centres, hist, label='Pairwisedistance data')
        plt.plot(bin_centres, hist_fit, label='Fitted data')
        plt.show()

    #
    # Fit data with gaussian mixter model allowing for a classifier between
    # fitted distribution of Burkitts pairwise distance of nuclei & DLBCL
    #
    def fitPairwiseDistGMM(pwDistBL, pwDistDL):
        #
        # simple application of gaussian mixture fit for pairwise distances
        #
        """
        gmm = GMM()
        gaussMixFit = gmm.fit(pwDistDL.flatten()[:,np.newaxis])
        print("GMM method depricated however worth a look")
        print("mean : %f, var : %f" % (gaussMixFit.means_[0, 0], gaussMixFit.covars_[0, 0]))
        """
        #
        # Use gaussian mixter model for training data labeled 0, 1 for
        # Burkitts, Diffuse Lymphoma respectively. 
        #
        DLlabel = [1 for i in pwDistDL]
        BLlabel = [0 for i in pwDistBL]
        DL_train = np.array([[i] for i in np.array(pwDistDL).flatten()])
        BL_train = np.array([[i] for i in np.array(pwDistBL).flatten()])
        X_train = [[i] for i in np.array([pwDistBL, pwDistDL]).flatten()]
        Y_train = [0, 1]
        estimators = dict((cov_type, GaussianMixture(n_components = 1, covariance_type=cov_type, max_iter=20, random_state=0)) for cov_type in ['tied'])
        #'spherical', 'tied',  'diag', 'full'
        n_estimators = len(estimators)
        for index, (name, estimator) in enumerate(estimators.items()):
            # Since we have class labels for the training data, we can
            # initialize the GMM parameters in a supervised manner.
            estimator.means_init = np.array([DL_train.mean(axis=0)])
            # now can add BL_flatten().mean(axis=0) and train on both data sets
            # Train the other parameters using the EM algorithm.
            estimator.fit(DL_train)
            print(estimator.predict(BL_train))
            print("Above is the gaussian mixture model fit to the DLBCL classifiying the Burkkitts Lymphoma, seems promising.")

    #
    # Fit gaussian to one set of pairwise distances for covariance/variance
    #
    #fitPairwiseDist(pwL2Dist)

    #
    # Perform Gaussian Mixed Model for pairwise distances between nuclei
    # in labeled Burkkitts and Diffuse large B-Cell lymphoma.
    #
    fitPairwiseDistGMM(BLpwDist, DLpwDist)

if args.analysis[0] == 'nDistDist':
    neighborDistanceDistribution(args.fBL[0], args.fDL[0])
elif args.analysis[0] == 'nucMorph':
    #
    # Function call for nuclei morphology abnormality classifier.
    #
    epsilon = args.e[0]
    filepath = args.f[0] #DLBCL = '../data/DLBCL results.xlsx' #BL = '../data/BL results.xlsx'
    nucleiMorphology(epsilon, filepath)
